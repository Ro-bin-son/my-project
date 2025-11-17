import os
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl import load_workbook
import time
import calendar


def addStudent(stnd_name, roll_num, students_dict):
    if roll_num in students_dict:
        print(f"Roll number {roll_num} already exists for student {students_dict[roll_num]}.")
    else:
        students_dict[roll_num] = stnd_name
        with open("Temprory.txt", "w") as file:
            json.dump(students_dict, file)
        file.close()
        print("Student Added Temprorily")


def allStudents(students_dict):
    print("----> Total Temprory Saved Students <----")
    for rollnum, student in students_dict.items():
        print(f"{rollnum} - {student}")


def addSubjects(subjt_name, subjects):
    subjects.add(subjt_name)
    if subjt_name in subjects:
        print("Subject Added Temprory")


def allSubjects(subjects):
    print("----> Total Temprory Saved Subjects <----")
    for subj_num, subject in enumerate(subjects, 1):
        print(f"{subj_num} - {subject}")


def createRegister(students_dict, subjects_set):
    print("1. Create Default Register")
    print("2. Create Dynamic Register")

    userChoice = int(input("Enter your choice : "))
    if (userChoice == 1):
        TIME = time.localtime()
        formated_time = time.strftime("%B-%Y", TIME)
        year = time.strftime("%Y", TIME)
        month = time.strftime("%m", TIME)
        total_days = calendar.monthrange(int(year), int(month))[1]

        wb = Workbook()
        sheet = wb.active
        sheet.title = formated_time
        col_headers = ['Roll no', 'Name']

        col_headers.extend([str(day) for day in range(1, total_days + 1)])
        for col_num, header in enumerate(col_headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

        if len(subjects_set) == 0:
            print("Please Add Subjects First")
        else:
            base_subjects = []
            for subject in subjects_set:
                base_subjects.append(subject)

        subjects = ['Subject', '']
        while len(subjects) - 2 <= total_days:
            subjects.extend(base_subjects)
        subjects = subjects[:total_days + 2]
        for col_num, subject in enumerate(subjects, 1):
            sheet.cell(row=2, column=col_num, value=subject)

        for row_num, (roll_no, name) in enumerate(students_dict.items(), start=3):
            sheet.cell(row=row_num, column=1, value=roll_no)
            sheet.cell(row=row_num, column=2, value=name)

        file_name = "Attendence_Register.xlsx"
        wb.save(file_name)
        print(f'Register create and saved as {file_name}')
    elif (userChoice == 2):
        save_as = input("Enter Register Name : ")
        Month = input("Enter a Month Name : ")
        Year = int(input("Enter Year : "))
        wb = Workbook()
        sheet = wb.active
        sheet.title = f"{Month}-{Year}"

        print("---> Write Data into Register <---")
        print("Default Target Cell 1 at Col-1, Row-1")
        print("Type exit for Exiting")
        loopRun = True
        while loopRun:
            print(
                "Select a Cell Example : A1, B2, C3....Z - A,B,C for Columns, 1,2,3,4.... for Rows. When Exceed Z then AA, BB, CC for Columns \n")
            print("""
            +----+----+----+----+----+----+----+----+----+----+----+
            | 1  | A1 | B1 | C1 | D1 | E1 | F1 | G1 | H1 | ...| Z1 |
            +----+----+----+----+----+----+----+----+----+----+----+
            | 2  | A2 | B2 | C2 | D2 | E2 | F2 | G2 | H2 | ...| Z2 |
            +----+----+----+----+----+----+----+----+----+----+----+
            | 3  | A3 | B3 | C3 | D3 | E3 | F3 | G3 | H3 | ...| Z3 |
            +----+----+----+----+----+----+----+----+----+----+----+
            | 4  | A4 | B4 | C4 | D4 | E4 | F4 | G4 | H4 | ...| Z4 |
            +----+----+----+----+----+----+----+----+----+----+----+
            | 5  | A5 | B5 | C5 | D5 | E5 | F5 | G5 | H5 | ...| Z5 |
            +----+----+----+----+----+----+----+----+----+----+----+
            """)

            cell = input("Enter a Cell Position : ")
            data = input("Enter a your data : ")
            if (cell == 'exit' or data == 'exit'):
                loopRun = False
            else:
                sheet[cell] = data
                wb.save(f"{save_as}.xlsx")


def readRegister():
    rgstr_name = input("Enter your Register Name : ")
    wb = load_workbook(f"{rgstr_name}.xlsx")
    sheet = wb.active
    for values in sheet.iter_rows(values_only=True):
        print(values)


def editRegister():
    rgstr_name = input("Enter your register Name : ")
    wb = load_workbook(f"./{rgstr_name}.xlsx")
    sheet = wb.active
    print("1. Deleting Rows & Columns")
    print("2. Merging & Unmerging Cells")
    print("3. Inserting Rows & Columns")
    print("4. Add Data to Specific Cell")

    match rgstr_name:
        case 1:
            print("1. Delete Rows")
            print("2. Delete Columns")
            select_opt = int(input("Select one of them : "))
            if (select_opt == 1):
                del_row = int(input("Enter a row number to Delete : "))
                if (1 <= del_row <= sheet.max_row):
                    sheet.delete_rows(del_row)
                    print("Row Deleted Successfully.")
                    wb.save(f"./{rgstr_name}.xlsx")
                else:
                    print('Row not found !!!')
            elif (select_opt == 2):
                del_cols = int(input("Enter a Col Position to Delete : "))
                if (1 <= del_cols <= sheet.max_column):
                    sheet.delete_cols(del_cols)
                    print("Column Deleted Successfully.")
                    wb.save(f"./{rgstr_name}.xlsx")
                else:
                    print('Column not found !!!')
            else:
                print("Option Not Exist !!!")
        case 2:
            print("Cells Position Example : A1:B1 to merge & Unmerge")
            print("1. Merge Cells")
            print("2. UnMerge Cells")
            select_opt = int(input("Select option one of them : "))
            if (select_opt == 1):
                select_cell1 = input("Select Cell 1")
                select_cell2 = input("Select Cell 2")

                cell1 = sheet[select_cell1]
                cell2 = sheet[select_cell2]

                if cell1.coordinate and cell2.coordinate:
                    sheet.merge_cells(f"{select_cell1}:{select_cell2}")
                    print("Cells Murged Succesfully.")
                else:
                    print("One or Both cells or not valid !!!")
            elif (select_opt == 2):
                select_cell1 = input("Select Cell 1")
                select_cell2 = input("Select Cell 2")

                cell1 = sheet[select_cell1]
                cell2 = sheet[select_cell2]

                if cell1.coordinate and cell2.coordinate:
                    sheet.unmerge_cells(f"{select_cell1}:{select_cell2}")
                    print("Cells Un-murged Succesfully.")
                else:
                    print("One or Both cells or not valid !!!")
            else:
                print("Option Not Exist !!!")
        case 3:
            print("Inserting Rows & Columns by Intering Positions.")
            print("1. Inserting Rows")
            print("2. Inserting Columns")
            userInput = int(input("Select option one of them : "))
            if (userInput == 1):
                insert_row = int(input("Enter a Position to Insert Row : "))
                if (insert_row > sheet.max_row + 1):
                    print(f"Total Rows exist in Sheet is : {sheet.max_row}")
                else:
                    sheet.insert_rows(insert_row)
                    print("Row Inserted Successfully.")
                    wb.save(f"./{rgstr_name}.xlsx")
            elif (userInput == 2):
                insert_col = int(input("Enter a Position to Insert Column : "))
                if (insert_col > sheet.max_row + 1):
                    print(f"Total Cols exist in Sheet is : {sheet.max_row}")
                else:
                    sheet.insert_cols(insert_col)
                    print("Column Inserted Successfully.")
                    wb.save(f"./{rgstr_name}.xlsx")
            else:
                print("Option Not Exist !!!")
        case 4:
            print("Insert Values to Specific Cell Example : A5, B5, C1...")
            input_cell = input("Enter a Cell Position : ")
            input_data = input("Enter a value to Insert : ")
            cell = sheet[input_cell]
            if cell.coordinate:
                sheet[input_cell] = input_data
                print("Value Inserting Successfully.")
                wb.save(f"./{rgstr_name}.xlsx")
            else:
                print("Cell Not found !!!")
        case _:
            print("Option Does not exist !!!")

    def check_attendence():
        print("Checking.... Total Attendence")
        register = input("Enter a name of Register : ")
        wb = load_workbook(f"{register}.xlsx")
        sheet = wb.active
        attendence_data = {}

        for row in sheet.iter_rows(min_row=3, values_only=True):
            roll_number = row[0]
            student_name = row[1]
            attendence_status = {}

            for idx, status in enumerate(row[2:], start=3):
                column_letter = openpyxl.utils.get_column_letter(idx)
                if status == 'y' or status == 'Y':
                    attendence_status[column_letter] = "Present"
                elif status == 'a' or status == 'A':
                    attendence_status[column_letter] = "Absent"
                else:
                    print("Empty Cells !!!")

            attendence_data[roll_number] = {
                "name": user_name,
                "status": attendence_status
            }

        for roll_number, data in attendence_data.items():
            print(f"Roll No :{roll_number}, Name :{data['name']}")
            for data, status in data['status'].items():
                print(f' Date Column {date}: {status}')


def main():
    print("======> Welcome to Attendence Management System <======")

    students = {}
    subjects = set()
    is_running = True

    while (is_running):
        print()
        print("1 - Add Students Temprory")
        print("2 - Add Subjects Temprory")
        print("3 - Show Total Temprory Students")
        print("4 - Show Total Temprory Subjects")
        print("5 - Create Attendence Register")
        print("6 - Read Class Attendence Register")
        print("7 - Edit Attendence Register")
        print("8 - Check Attedence in the end of Month")
        print("9 - Exit")
        print()

        user_selection = int(input("Enter a Number : "))

        match user_selection:
            case 1:
                case1 = True
                while (case1):
                    student_name = input("Enter a Student Name : ")
                    if (student_name == "exit"):
                        case1 = False
                        break
                    roll_num = int(input("Enter Roll Number : "))
                    addStudent(student_name, roll_num, students)
                    print("Type exit for Exiting")
                    print("------------------------------")
            case 2:
                case2 = True
                while case2:
                    subject_name = input("Enter a Subject Name : ")
                    if (subject_name == 'exit'):
                        case2 = False
                        break
                    addSubjects(subject_name, subjects)
                    print("Type exit of Exiting")
                    print("------------------------------")
            case 3:
                allStudents(students)
            case 4:
                allSubjects(subjects)
            case 5:
                createRegister(students, subjects)
            case 6:
                readRegister()
            case 7:
                editRegister()
            case 8:
                check_attendence()
            case 9:
                is_running = False
            case _:
                print("Invalid Input")


if __name__ == "__main__":
    main()